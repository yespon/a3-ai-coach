from dataclasses import dataclass
from io import BytesIO
from typing import Any
import uuid

from openpyxl import Workbook, load_workbook

from app.services.whitelist_service import MAX_WHITELIST_ROWS

ROLE_LABELS = {"管理员": "admin", "教练": "coach", "学员": "student", "admin": "admin", "coach": "coach", "student": "student"}
ENABLED_LABELS = {"启用": True, "禁用": False, "true": True, "false": False, "是": True, "否": False}
MANAGED_USER_TEMPLATE_HEADERS = ["工号", "姓名", "邮箱", "一级部门", "主角色", "兼任教练", "所属教练工号", "启用状态"]


@dataclass
class ManagedUserParseResult:
    rows: list[dict[str, Any]]
    errors: list[dict[str, Any]]


def _cell_text(value) -> str:
    return str(value).strip() if value is not None else ""


def normalize_employee_no(value: str) -> str:
    return value.strip()


def is_effective_coach(primary_role: str, is_coach: bool) -> bool:
    return primary_role == "coach" or (primary_role == "admin" and is_coach)


def normalize_managed_user_role(
    primary_role: str | None, is_coach: bool, coach_id: str | uuid.UUID | None
):
    role = ROLE_LABELS.get((primary_role or "学员").strip())
    if role is None:
        raise ValueError("主角色必须是管理员、教练或学员")
    if role == "coach":
        return {"primary_role": "coach", "is_coach": True, "coach_id": None}
    if role == "admin":
        return {"primary_role": "admin", "is_coach": bool(is_coach), "coach_id": None}
    return {"primary_role": "student", "is_coach": False, "coach_id": coach_id}


def protect_system_admin_patch(employee_no: str, admin_employee_nos: set[str], requested: dict[str, Any]) -> dict[str, Any]:
    if employee_no not in admin_employee_nos:
        return requested
    protected = dict(requested)
    protected["enabled"] = True
    protected["primary_role"] = "admin"
    return protected


def build_managed_user_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "用户管理"
    ws.append(MANAGED_USER_TEMPLATE_HEADERS)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header_indexes(header: list[str]) -> dict[str, int] | None:
    try:
        return {name: header.index(name) for name in MANAGED_USER_TEMPLATE_HEADERS}
    except ValueError:
        return None


def _parse_bool(value: str, default: bool) -> bool | None:
    if not value:
        return default
    lowered = value.lower()
    return ENABLED_LABELS.get(lowered) if lowered in ENABLED_LABELS else ENABLED_LABELS.get(value)


def parse_managed_user_excel(raw: bytes) -> ManagedUserParseResult:
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    header = [_cell_text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    indexes = _header_indexes(header)
    if indexes is None:
        return ManagedUserParseResult(rows=[], errors=[{"row": 1, "reason": "表头必须包含工号、姓名、邮箱、一级部门、主角色、兼任教练、所属教练工号、启用状态"}])

    latest: dict[str, dict[str, Any]] = {}
    row_numbers: dict[str, int] = {}
    errors: list[dict[str, Any]] = []
    for row_no, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row_no - 1 > MAX_WHITELIST_ROWS:
            errors.append({"row": row_no, "reason": f"超过最大导入行数 {MAX_WHITELIST_ROWS}"})
            break
        employee_no = normalize_employee_no(_cell_text(row[indexes["工号"]].value if indexes["工号"] < len(row) else None))
        if not employee_no:
            errors.append({"row": row_no, "reason": "工号为空"})
            continue
        role_text = _cell_text(row[indexes["主角色"]].value if indexes["主角色"] < len(row) else None) or "学员"
        coach_flag_text = _cell_text(row[indexes["兼任教练"]].value if indexes["兼任教练"] < len(row) else None)
        coach_flag = _parse_bool(coach_flag_text, False)
        if coach_flag is None:
            errors.append({"row": row_no, "reason": "兼任教练必须是是或否"})
            continue
        try:
            normalized = normalize_managed_user_role(role_text, coach_flag, None)
        except ValueError as exc:
            errors.append({"row": row_no, "reason": str(exc)})
            continue
        if normalized["primary_role"] == "student" and coach_flag:
            errors.append({"row": row_no, "reason": "学员不能兼任教练"})
            continue
        enabled_text = _cell_text(row[indexes["启用状态"]].value if indexes["启用状态"] < len(row) else None)
        enabled = _parse_bool(enabled_text, True)
        if enabled is None:
            errors.append({"row": row_no, "reason": "启用状态必须是启用或禁用"})
            continue
        latest[employee_no] = {
            "employee_no": employee_no,
            "name": _cell_text(row[indexes["姓名"]].value if indexes["姓名"] < len(row) else None) or None,
            "email": _cell_text(row[indexes["邮箱"]].value if indexes["邮箱"] < len(row) else None) or None,
            "department_level1": _cell_text(row[indexes["一级部门"]].value if indexes["一级部门"] < len(row) else None) or None,
            "primary_role": normalized["primary_role"],
            "is_coach": normalized["is_coach"],
            "coach_employee_no": _cell_text(row[indexes["所属教练工号"]].value if indexes["所属教练工号"] < len(row) else None) or None,
            "enabled": enabled,
        }
        row_numbers[employee_no] = row_no
    rows = []
    for employee_no, payload in latest.items():
        payload["_row"] = row_numbers[employee_no]
        rows.append(payload)
    for payload in rows:
        payload.pop("_row")
    return ManagedUserParseResult(rows=rows, errors=errors)


def resolve_import_coach_links(rows: list[dict[str, Any]], existing_coach_employee_nos: set[str]) -> list[dict[str, Any]]:
    batch_coaches = {row["employee_no"] for row in rows if is_effective_coach(row["primary_role"], row["is_coach"])}
    valid_coaches = existing_coach_employee_nos | batch_coaches
    errors: list[dict[str, Any]] = []
    for idx, row in enumerate(rows, start=2):
        coach_employee_no = row.get("coach_employee_no")
        if row["primary_role"] == "student" and coach_employee_no and coach_employee_no not in valid_coaches:
            errors.append({"row": idx, "reason": "所属教练工号不存在或不是教练"})
    return errors
