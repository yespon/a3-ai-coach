"""SSO whitelist service — query, Excel import/export, upsert."""

from dataclasses import dataclass
from io import BytesIO
from typing import Any
import uuid

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.db_models import SsoUserWhitelistDB

WHITELIST_DENY_MESSAGE = "当前账号未开通 A3工作法AI教练 访问权限，请联系管理员开通。"
MAX_WHITELIST_UPLOAD_BYTES = 5 * 1024 * 1024
MAX_WHITELIST_ROWS = 50_000


@dataclass
class ParseResult:
    rows: list[dict[str, str | None]]
    errors: list[dict[str, Any]]


def _cell_text(value) -> str:
    return str(value).strip() if value is not None else ""


def parse_whitelist_excel(raw: bytes) -> ParseResult:
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    header = [_cell_text(c.value) for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        emp_idx = header.index("工号")
        email_idx = header.index("邮箱")
    except ValueError:
        return ParseResult(rows=[], errors=[{"row": 1, "reason": "表头必须包含工号、邮箱"}])
    latest: dict[str, str | None] = {}
    errors: list[dict[str, Any]] = []
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if idx - 1 > MAX_WHITELIST_ROWS:
            errors.append({"row": idx, "reason": f"超过最大导入行数 {MAX_WHITELIST_ROWS}"})
            break
        employee_no = normalize_employee_no(_cell_text(row[emp_idx].value if emp_idx < len(row) else None))
        email = _cell_text(row[email_idx].value if email_idx < len(row) else None) or None
        if not employee_no:
            errors.append({"row": idx, "reason": "工号为空"})
            continue
        latest[employee_no] = email
    rows = [{"employee_no": k, "email": v} for k, v in latest.items()]
    return ParseResult(rows=rows, errors=errors)


def normalize_employee_no(value: str) -> str:
    return value.strip()


def build_whitelist_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "白名单"
    ws.append(["工号", "邮箱"])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def is_employee_allowed(db: AsyncSession, employee_no: str) -> bool:
    result = await db.execute(select(SsoUserWhitelistDB).where(
        SsoUserWhitelistDB.employee_no == employee_no,
        SsoUserWhitelistDB.enabled == True,
    ))
    return result.scalar_one_or_none() is not None


async def upsert_whitelist_entry(
    db: AsyncSession, employee_no: str, email: str | None,
    source: str, created_by: uuid.UUID | None,
):
    result = await db.execute(
        select(SsoUserWhitelistDB).where(SsoUserWhitelistDB.employee_no == employee_no)
    )
    entry = result.scalar_one_or_none()
    if entry is None:
        entry = SsoUserWhitelistDB(
            employee_no=employee_no, email=email, enabled=True,
            source=source, created_by=created_by,
        )
        db.add(entry)
        created = True
    else:
        entry.email = email
        entry.source = source
        created = False
    return entry, created
