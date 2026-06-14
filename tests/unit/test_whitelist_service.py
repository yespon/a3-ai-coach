from io import BytesIO
from openpyxl import Workbook, load_workbook

import pytest

from app.models.db_models import SsoUserWhitelistDB
from app.services.whitelist_service import parse_whitelist_excel, build_whitelist_template, upsert_whitelist_entry


def _xlsx(rows):
    wb = Workbook(); ws = wb.active
    for row in rows: ws.append(row)
    buf = BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()


def test_parse_whitelist_excel_valid_rows():
    data = _xlsx([["工号", "邮箱"], [" 1001 ", "a@example.com"], ["1002", None]])
    result = parse_whitelist_excel(data)
    assert result.rows == [{"employee_no": "1001", "email": "a@example.com"}, {"employee_no": "1002", "email": None}]
    assert result.errors == []


def test_parse_whitelist_excel_reports_empty_employee_no():
    data = _xlsx([["工号", "邮箱"], [None, "a@example.com"]])
    result = parse_whitelist_excel(data)
    assert result.rows == []
    assert result.errors == [{"row": 2, "reason": "工号为空"}]


def test_template_contains_expected_headers():
    wb = load_workbook(BytesIO(build_whitelist_template()))
    assert [cell.value for cell in wb.active[1]] == ["工号", "邮箱"]


def test_parse_whitelist_excel_enforces_row_limit(monkeypatch):
    from app.services import whitelist_service
    monkeypatch.setattr(whitelist_service, "MAX_WHITELIST_ROWS", 1)
    data = _xlsx([["工号", "邮箱"], ["1001", "a@example.com"], ["1002", "b@example.com"]])
    result = parse_whitelist_excel(data)
    assert result.rows == [{"employee_no": "1001", "email": "a@example.com"}]
    assert result.errors == [{"row": 3, "reason": "超过最大导入行数 1"}]


class FakeResult:
    def __init__(self, entry): self.entry = entry
    def scalar_one_or_none(self): return self.entry


class FakeSession:
    def __init__(self, entry): self.entry = entry
    async def execute(self, stmt): return FakeResult(self.entry)
    def add(self, obj): self.entry = obj


@pytest.mark.asyncio
async def test_upsert_whitelist_entry_preserves_disabled_existing_entry():
    entry = SsoUserWhitelistDB(employee_no="1001", email="old@example.com", enabled=False, source="manual")
    db = FakeSession(entry)
    updated, created = await upsert_whitelist_entry(db, "1001", "new@example.com", "excel", None)
    assert created is False
    assert updated.email == "new@example.com"
    assert updated.enabled is False
    assert updated.source == "excel"
