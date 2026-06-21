from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.models.db_models import ManagedUserDB
from app.services.managed_user_service import (
    MANAGED_USER_TEMPLATE_HEADERS,
    build_managed_user_template,
    ensure_managed_user_allowed,
    is_effective_coach,
    normalize_managed_user_role,
    parse_managed_user_excel,
    protect_system_admin_patch,
    resolve_import_coach_links,
    upsert_managed_user,
)


def test_normalize_student_role_clears_coach_fields():
    normalized = normalize_managed_user_role("学员", True, "coach-id")
    assert normalized == {"primary_role": "student", "is_coach": False, "coach_id": "coach-id"}


def test_normalize_coach_role_forces_is_coach_and_clears_coach_id():
    normalized = normalize_managed_user_role("岗位负责人", False, "coach-id")
    assert normalized == {"primary_role": "coach", "is_coach": True, "coach_id": None}


def test_normalize_admin_role_keeps_admin_coach_capability_and_clears_coach_id():
    normalized = normalize_managed_user_role("管理员", True, "coach-id")
    assert normalized == {"primary_role": "admin", "is_coach": True, "coach_id": None}


def test_effective_coach_includes_coach_and_admin_coach():
    assert is_effective_coach("coach", False) is True
    assert is_effective_coach("admin", True) is True
    assert is_effective_coach("admin", False) is False
    assert is_effective_coach("student", False) is False


def test_system_admin_patch_cannot_disable_or_downgrade():
    protected = protect_system_admin_patch(
        employee_no="1001",
        admin_employee_nos={"1001"},
        requested={"enabled": False, "primary_role": "student", "is_coach": False},
    )
    assert protected["enabled"] is True
    assert protected["primary_role"] == "admin"
    assert protected["is_coach"] is False


def test_managed_user_template_headers_are_in_confirmed_order():
    wb = load_workbook(BytesIO(build_managed_user_template()))
    assert [cell.value for cell in wb.active[1]] == MANAGED_USER_TEMPLATE_HEADERS
    assert MANAGED_USER_TEMPLATE_HEADERS == [
        "工号",
        "姓名",
        "邮箱",
        "一级部门",
        "主角色",
        "兼任负责人",
        "所属负责人工号",
        "启用状态",
    ]


def _xlsx(rows):
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_parse_managed_user_excel_defaults_student_and_enabled():
    data = _xlsx([
        ["工号", "姓名", "邮箱", "一级部门", "主角色", "兼任负责人", "所属负责人工号", "启用状态"],
        [" 1001 ", "张三", "a@example.com", "研发", None, None, None, None],
    ])
    result = parse_managed_user_excel(data)
    assert result.errors == []
    assert result.rows == [{
        "employee_no": "1001",
        "name": "张三",
        "email": "a@example.com",
        "department_level1": "研发",
        "primary_role": "student",
        "is_coach": False,
        "coach_employee_no": None,
        "enabled": True,
        "row": 2,
    }]


def test_parse_managed_user_excel_keeps_department_after_email():
    data = _xlsx([
        ["工号", "姓名", "邮箱", "一级部门", "主角色", "兼任负责人", "所属负责人工号", "启用状态"],
        ["2001", "李教练", "coach@example.com", "销售", "岗位负责人", "否", None, "启用"],
    ])
    result = parse_managed_user_excel(data)
    assert result.rows[0]["department_level1"] == "销售"
    assert result.rows[0]["primary_role"] == "coach"
    assert result.rows[0]["is_coach"] is True


def test_parse_managed_user_excel_warns_duplicate_employee_no_and_keeps_last_row():
    data = _xlsx([
        ["工号", "姓名", "邮箱", "一级部门", "主角色", "兼任负责人", "所属负责人工号", "启用状态"],
        ["1001", "张三", "first@example.com", "研发", "学员", "否", None, "启用"],
        ["1001", "张三更新", "last@example.com", "销售", "岗位负责人", "否", None, "禁用"],
    ])

    result = parse_managed_user_excel(data)

    assert result.errors == [{"row": 3, "reason": "工号重复，已使用最后一条记录"}]
    assert result.rows == [{
        "employee_no": "1001",
        "name": "张三更新",
        "email": "last@example.com",
        "department_level1": "销售",
        "primary_role": "coach",
        "is_coach": True,
        "coach_employee_no": None,
        "enabled": False,
        "row": 3,
    }]


def test_missing_coach_after_duplicate_reports_kept_original_row_number():
    data = _xlsx([
        ["工号", "姓名", "邮箱", "一级部门", "主角色", "兼任负责人", "所属负责人工号", "启用状态"],
        ["1001", "张三", None, None, "学员", "否", None, "启用"],
        ["1001", "张三更新", None, None, "学员", "否", "9999", "启用"],
    ])

    parse_result = parse_managed_user_excel(data)
    errors = resolve_import_coach_links(parse_result.rows, existing_coach_employee_nos=set())

    assert parse_result.errors == [{"row": 3, "reason": "工号重复，已使用最后一条记录"}]
    assert parse_result.rows[0]["row"] == 3
    assert errors == [{"row": 3, "reason": "所属负责人工号不存在或不是岗位负责人"}]

def test_parse_managed_user_excel_rejects_invalid_student_coach_flag():
    data = _xlsx([
        ["工号", "姓名", "邮箱", "一级部门", "主角色", "兼任负责人", "所属负责人工号", "启用状态"],
        ["1001", "张三", None, None, "学员", "是", None, "启用"],
    ])
    result = parse_managed_user_excel(data)
    assert result.rows == []
    assert result.errors == [{"row": 2, "reason": "学员不能兼任负责人"}]


def test_resolve_import_coach_links_accepts_same_batch_coach():
    rows = [
        {"employee_no": "2001", "primary_role": "coach", "is_coach": True, "coach_employee_no": None},
        {"employee_no": "1001", "primary_role": "student", "is_coach": False, "coach_employee_no": "2001"},
    ]
    errors = resolve_import_coach_links(rows, existing_coach_employee_nos=set())
    assert errors == []


def test_resolve_import_coach_links_rejects_missing_coach():
    rows = [
        {"employee_no": "1001", "primary_role": "student", "is_coach": False, "coach_employee_no": "9999"},
    ]
    errors = resolve_import_coach_links(rows, existing_coach_employee_nos=set())
    assert errors == [{"row": 2, "reason": "所属负责人工号不存在或不是岗位负责人"}]


class FakeScalarResult:
    def __init__(self, value):
        self.value = value

    def scalar_one_or_none(self):
        return self.value

    def scalars(self):
        return self

    def all(self):
        return self.value if isinstance(self.value, list) else []


class FakeDb:
    def __init__(self, execute_values=None):
        self.execute_values = list(execute_values or [])
        self.added = []
        self.committed = False
        self.refreshed = []

    async def execute(self, stmt):
        value = self.execute_values.pop(0) if self.execute_values else None
        return FakeScalarResult(value)

    def add(self, obj):
        self.added.append(obj)

    async def commit(self):
        self.committed = True

    async def refresh(self, obj):
        self.refreshed.append(obj)


@pytest.mark.asyncio
async def test_ensure_managed_user_allowed_rejects_missing_non_system_admin():
    db = FakeDb([None])
    with pytest.raises(PermissionError) as exc:
        await ensure_managed_user_allowed(db, "1001", False)
    assert str(exc.value) == "当前账号未开通 A3工作法AI教练 访问权限，请联系管理员开通。"


@pytest.mark.asyncio
async def test_ensure_managed_user_allowed_allows_enabled_profile():
    profile = ManagedUserDB(employee_no="1001", enabled=True, primary_role="student", is_coach=False)
    db = FakeDb([profile])
    assert await ensure_managed_user_allowed(db, "1001", False) is profile


@pytest.mark.asyncio
async def test_ensure_managed_user_allowed_creates_missing_system_admin_profile():
    db = FakeDb([None])
    profile = await ensure_managed_user_allowed(db, "9999", True)
    assert profile.employee_no == "9999"
    assert profile.primary_role == "admin"
    assert profile.enabled is True
    assert profile in db.added
    assert db.committed is True


@pytest.mark.asyncio
async def test_upsert_managed_user_updates_existing_profile_without_disabling_system_admin():
    profile = ManagedUserDB(employee_no="9999", enabled=True, primary_role="admin", is_coach=False)
    db = FakeDb([profile])
    updated, created = await upsert_managed_user(
        db,
        {
            "employee_no": "9999",
            "name": "系统管理员",
            "email": "admin@example.com",
            "department_level1": "总部",
            "primary_role": "student",
            "is_coach": False,
            "coach_id": None,
            "enabled": False,
        },
        source="manual",
        created_by=None,
        admin_employee_nos={"9999"},
    )
    assert created is False
    assert updated.primary_role == "admin"
    assert updated.enabled is True
    assert updated.name == "系统管理员"
