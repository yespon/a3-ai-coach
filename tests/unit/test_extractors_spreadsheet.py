# T3-2: spreadsheet merged-cell extraction should preserve repeated values.

from io import BytesIO

from openpyxl import Workbook

from app.extractors.spreadsheet import _extract_xlsx_text


def test_extract_xlsx_fills_merged_cells():
    wb = Workbook()
    ws = wb.active
    ws.title = "任务表"
    ws["A1"] = "目的:"
    ws["B1"] = "提升团队协作"
    ws["A2"] = "成果:"
    ws["B2"] = "预算可控"

    ws["A4"] = "核心任务"
    ws.merge_cells("A4:A5")
    ws["B4"] = "拆分目标"
    ws["B5"] = "执行追踪"

    buf = BytesIO()
    wb.save(buf)

    text = _extract_xlsx_text(buf.getvalue())

    assert "[Sheet] 任务表" in text
    assert "任务目的：" in text
    assert "任务成果（预算、交期、完成度）：" in text
    # Merged value from A4 should appear on both logical rows after fill.
    assert "核心任务\t拆分目标" in text
    assert "核心任务\t执行追踪" in text
